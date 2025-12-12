"""
JSON to Excel Converter for ACU Course Data
============================================
Converts the extracted ACU course JSON data into a structured Excel file
with multiple sheets for easy readability.
"""

import json
import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def load_json_data(json_path: str) -> dict:
    """Load JSON data from file."""
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def flatten_course_data(data: dict) -> list:
    """Flatten nested course data into a list of dictionaries."""
    courses = []

    if "courses_by_study_area" in data:
        for study_area, course_list in data["courses_by_study_area"].items():
            for course in course_list:
                flat_course = {
                    "Course Name": course.get("course_name", "Not Found"),
                    "Subject Area": study_area,
                    "Department": course.get("department", "Not Found"),
                    "Study Level": course.get("study_level", "Not Found"),
                    "Course Type": course.get("course_type", "Not Found"),
                    "Duration": course.get("duration", "Not Found"),
                    "Full-Time": "Yes" if course.get("study_mode", {}).get("full_time") else "No",
                    "Part-Time": "Yes" if course.get("study_mode", {}).get("part_time") else "No",
                    "Online": "Yes" if course.get("delivery_mode", {}).get("online") else "No",
                    "On-Campus": "Yes" if course.get("delivery_mode", {}).get("on_campus") else "No",
                    "Blended": "Yes" if course.get("delivery_mode", {}).get("blended") else "No",
                    "Campuses": ", ".join(course.get("campuses", [])),
                    "Domestic Fee (CSP)": course.get("fees", {}).get("domestic_csp", "Not Found"),
                    "Domestic Fee (Full)": course.get("fees", {}).get("domestic_fee_paying", "Not Found"),
                    "International Fee": course.get("fees", {}).get("international", "Not Found"),
                    "ATAR Requirement": course.get("atar_requirement", "Not Found"),
                    "Intake Periods": ", ".join(course.get("intake_periods", [])),
                    "Description": course.get("description", "Not Found")[:300] + "..." if len(course.get("description", "")) > 300 else course.get("description", "Not Found"),
                    "URL": course.get("url", "Not Found"),
                }
                courses.append(flat_course)

    return courses


def create_summary_df(data: dict) -> pd.DataFrame:
    """Create summary dataframe."""
    summary = data.get("summary", {})

    summary_data = [
        ["University", data.get("university", "N/A")],
        ["Website", data.get("website", "N/A")],
        ["Extraction Date", data.get("extraction_date", "N/A")],
        ["Extraction Time", data.get("extraction_time", "N/A")],
        ["", ""],
        ["Total Courses", summary.get("total_courses", 0)],
        ["Online Courses Available", summary.get("online_available", 0)],
        ["Part-Time Available", summary.get("part_time_available", 0)],
    ]

    return pd.DataFrame(summary_data, columns=["Metric", "Value"])


def create_by_study_level_df(data: dict) -> pd.DataFrame:
    """Create study level breakdown dataframe."""
    summary = data.get("summary", {})
    by_level = summary.get("by_study_level", {})

    level_data = [[level, count] for level, count in by_level.items()]
    return pd.DataFrame(level_data, columns=["Study Level", "Count"])


def create_by_course_type_df(data: dict) -> pd.DataFrame:
    """Create course type breakdown dataframe."""
    summary = data.get("summary", {})
    by_type = summary.get("by_course_type", {})

    type_data = [[ctype, count] for ctype, count in by_type.items()]
    return pd.DataFrame(type_data, columns=["Course Type", "Count"])


def create_by_subject_area_df(data: dict) -> pd.DataFrame:
    """Create subject area breakdown dataframe."""
    summary = data.get("summary", {})
    by_area = summary.get("by_subject_area", {})

    area_data = [[area, count] for area, count in sorted(by_area.items())]
    return pd.DataFrame(area_data, columns=["Subject Area", "Count"])


def create_by_campus_df(data: dict) -> pd.DataFrame:
    """Create campus breakdown dataframe."""
    summary = data.get("summary", {})
    by_campus = summary.get("by_campus", {})

    campus_data = [[campus, count] for campus, count in sorted(by_campus.items())]
    return pd.DataFrame(campus_data, columns=["Campus (City)", "Courses Available"])


def style_worksheet(ws, df, title: str, is_main_sheet: bool = False):
    """Apply styling to worksheet."""
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Add title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14, color="2E75B6")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data starting from row 3
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = cell_border

            if r_idx == 3:  # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            else:
                cell.alignment = cell_alignment
                # Alternate row coloring
                if r_idx % 2 == 0:
                    cell.fill = alt_fill

    # Auto-adjust column widths
    for col_idx, column in enumerate(df.columns, start=1):
        max_length = len(str(column))
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, min(len(str(cell.value)), 50))
                except:
                    pass

        # Set column width
        if is_main_sheet:
            if column == "Description":
                adjusted_width = 60
            elif column == "URL":
                adjusted_width = 45
            elif column == "Course Name":
                adjusted_width = 45
            elif column in ["Campuses", "Intake Periods"]:
                adjusted_width = 35
            else:
                adjusted_width = min(max_length + 2, 30)
        else:
            adjusted_width = min(max_length + 4, 40)

        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Set row height for header
    ws.row_dimensions[3].height = 25

    # Freeze panes (freeze header row)
    ws.freeze_panes = ws['A4']


def create_subject_area_sheets(wb, data: dict):
    """Create individual sheets for each subject area."""
    if "courses_by_study_area" not in data:
        return

    for study_area, courses in data["courses_by_study_area"].items():
        # Create safe sheet name (max 31 chars, no special chars)
        sheet_name = study_area.replace("&", "and").replace("/", "-")[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Flatten courses for this area
        flat_courses = []
        for course in courses:
            flat_course = {
                "Course Name": course.get("course_name", "Not Found"),
                "Course Type": course.get("course_type", "Not Found"),
                "Study Level": course.get("study_level", "Not Found"),
                "Duration": course.get("duration", "Not Found"),
                "Full-Time": "Yes" if course.get("study_mode", {}).get("full_time") else "No",
                "Part-Time": "Yes" if course.get("study_mode", {}).get("part_time") else "No",
                "Online": "Yes" if course.get("delivery_mode", {}).get("online") else "No",
                "On-Campus": "Yes" if course.get("delivery_mode", {}).get("on_campus") else "No",
                "Domestic Fee (CSP)": course.get("fees", {}).get("domestic_csp", "Not Found"),
                "ATAR": course.get("atar_requirement", "Not Found"),
                "Campuses": ", ".join(course.get("campuses", [])),
            }
            flat_courses.append(flat_course)

        df = pd.DataFrame(flat_courses)
        style_worksheet(ws, df, f"{study_area} Courses", is_main_sheet=False)


def json_to_excel(json_path: str, excel_path: str):
    """Convert JSON data to structured Excel file."""
    print(f"Loading JSON data from: {json_path}")
    data = load_json_data(json_path)

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # 1. Summary Sheet
    print("Creating Summary sheet...")
    ws_summary = wb.create_sheet(title="Summary")
    summary_df = create_summary_df(data)
    style_worksheet(ws_summary, summary_df, "ACU Course Data - Summary", is_main_sheet=False)

    # Add additional summary tables
    # Study Level breakdown
    level_df = create_by_study_level_df(data)
    start_row = len(summary_df) + 6
    ws_summary.cell(row=start_row, column=1, value="By Study Level").font = Font(bold=True, size=12)
    for r_idx, row in enumerate(dataframe_to_rows(level_df, index=False, header=True), start=start_row + 1):
        for c_idx, value in enumerate(row, start=1):
            ws_summary.cell(row=r_idx, column=c_idx, value=value)

    # Course Type breakdown
    type_df = create_by_course_type_df(data)
    start_row = start_row + len(level_df) + 4
    ws_summary.cell(row=start_row, column=1, value="By Course Type").font = Font(bold=True, size=12)
    for r_idx, row in enumerate(dataframe_to_rows(type_df, index=False, header=True), start=start_row + 1):
        for c_idx, value in enumerate(row, start=1):
            ws_summary.cell(row=r_idx, column=c_idx, value=value)

    # 2. All Courses Sheet
    print("Creating All Courses sheet...")
    ws_all = wb.create_sheet(title="All Courses")
    all_courses = flatten_course_data(data)
    all_courses_df = pd.DataFrame(all_courses)
    style_worksheet(ws_all, all_courses_df, "All ACU Courses - Complete List", is_main_sheet=True)

    # 3. By Subject Area Sheet
    print("Creating By Subject Area sheet...")
    ws_area = wb.create_sheet(title="By Subject Area")
    area_df = create_by_subject_area_df(data)
    style_worksheet(ws_area, area_df, "Courses by Subject Area", is_main_sheet=False)

    # 4. By Campus Sheet
    print("Creating By Campus sheet...")
    ws_campus = wb.create_sheet(title="By Campus")
    campus_df = create_by_campus_df(data)
    style_worksheet(ws_campus, campus_df, "Courses by Campus Location", is_main_sheet=False)

    # 5. Individual Subject Area Sheets
    print("Creating individual subject area sheets...")
    create_subject_area_sheets(wb, data)

    # Save workbook
    print(f"\nSaving Excel file to: {excel_path}")
    wb.save(excel_path)
    print("Excel file created successfully!")

    # Print summary
    print("\n" + "=" * 60)
    print("EXCEL FILE STRUCTURE")
    print("=" * 60)
    print("\nSheets created:")
    for idx, sheet in enumerate(wb.sheetnames, 1):
        print(f"  {idx}. {sheet}")
    print(f"\nTotal courses: {len(all_courses)}")
    print("=" * 60)


def main():
    """Main function."""
    # Paths
    script_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(script_dir, "output", "acu", "acu_courses_by_study_area.json")
    excel_path = os.path.join(script_dir, "output", "acu", "acu_courses_data.xlsx")

    # Check if JSON file exists
    if not os.path.exists(json_path):
        print(f"Error: JSON file not found at {json_path}")
        print("Please run acu_crawler.py first to generate the JSON data.")
        return

    # Convert to Excel
    json_to_excel(json_path, excel_path)


if __name__ == "__main__":
    main()
